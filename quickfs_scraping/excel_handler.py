import pandas as pd
import os
import pymsgbox
import openpyxl
from pandas import ExcelWriter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import colors, Font

from quickfs_scraping.general import check_validity_output_file, get_sheet_name


def excel_to_dataframe(excel_output_path, source='Web Scraping'):
    excel_file = pd.ExcelFile(excel_output_path, engine='openpyxl')
    sheet_name = get_sheet_name(scraping_method=source)

    return pd.read_excel(excel_file, sheet_name=sheet_name)


def dataframe_to_excel(excel_output_path, df, ticker, source='API', bool_batch=False):
    # Get first a valid sheet name
    if source != 'API' and source != 'Web Scraping':
        sheet_name = get_sheet_name(bool_scraping=False)
    else:
        sheet_name = get_sheet_name(scraping_method=source)

    # Check if output file is still valid
    if check_validity_output_file(excel_output_path):
        delete_excel_sheet_from_wb(excel_output_path, sheet_name)

        # Create pandas excel writer object, and include new table without deleting other sheets
        wb = openpyxl.load_workbook(filename=excel_output_path)
        with ExcelWriter(excel_output_path, engine='openpyxl') as writer:
            writer.book = wb
            df.to_excel(excel_writer=writer, sheet_name=sheet_name, index=False)
            writer.save()
        wb.close()
    else:
        # Just create a new workbook to dump the data
        with ExcelWriter(excel_output_path, engine='openpyxl') as writer:
            df.to_excel(excel_writer=writer, sheet_name=sheet_name, index=False)
            writer.save()

    # Change exported data in excel to a table format
    excel_data_to_table(excel_output_path, sheet_name, df)

    # Success message
    if not bool_batch:
        user_answer = pymsgbox.confirm(f'The financial excel file for {ticker} has been successfully generated. '
                                       f'Do you want to open the excel file?', 'Success!', buttons=['OK', 'Cancel'])

        if user_answer == 'OK':
            os.startfile(excel_output_path)


def delete_excel_sheet_from_wb(excel_path, sheet_name):
    # Delete sheet if already exists
    wb = openpyxl.load_workbook(filename=excel_path)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    wb.save(excel_path)
    wb.close()


def change_sheet_names(excel_file, sheet_name):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb[sheet_name]
    ws.title = sheet_name + '_scraped'
    wb.save(excel_file)
    wb.close()


def change_table_name(excel_file, sheet_name, old_table_name, new_table_name):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb[sheet_name]

    my_table = ws.tables[old_table_name]
    my_table.displayName = new_table_name

    wb.save(excel_file)
    wb.close()


def excel_data_to_table(excel_output_path, sheet_name, df):
    # Convert data in excel file to table format
    wb = openpyxl.load_workbook(filename=excel_output_path)

    # Create table
    tab = Table(displayName=sheet_name, ref=f'A1:{chr(len(df.columns) + 64)}{len(df) + 1}')

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style

    # Change header font color
    # ws = wb[sheet_name]
    # header_col = ws.row_dimensions[1]
    # header_col.font = Font(color=colors.WHITE)

    # Add table to sheet and save workbook
    wb[sheet_name].add_table(tab)
    wb.save(excel_output_path)
    wb.close()


def excel_sheet_exists(excel_output_path, source='Web Scraping'):
    if check_validity_output_file(excel_output_path):
        wb = openpyxl.load_workbook(excel_output_path, read_only=True)
        sheet_name = get_sheet_name(scraping_method=source)
        if sheet_name in wb.sheetnames:
            wb.close()
            return True
        else:
            wb.close()
            return False
    else:
        return False
